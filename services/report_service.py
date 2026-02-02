"""
車行寶 CRM v5.1 - 報表服務模組
北斗七星文創數位 × 織明

功能：
1. 日報/週報/月報自動生成
2. 業績排行榜
3. Excel 匯出
4. 定時報表推送
"""
import io
import json
from datetime import datetime, timedelta
from typing import Dict, List, Optional
from models import get_connection

# 嘗試導入 openpyxl（Excel 處理）
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ============================================================
# 1. 日報
# ============================================================

def generate_daily_report(db_path: str, date: Optional[str] = None) -> Dict:
    """生成日報
    
    Args:
        db_path: 資料庫路徑
        date: 日期 (YYYY-MM-DD)，預設今天
    
    Returns:
        完整日報數據
    """
    if not date:
        date = datetime.now().strftime('%Y-%m-%d')
    
    conn = get_connection(db_path)
    c = conn.cursor()
    
    report = {
        'type': 'daily',
        'date': date,
        'generated_at': datetime.now().isoformat()
    }
    
    # 1. 銷售統計
    c.execute('''
        SELECT COUNT(*) as count, 
               SUM(amount) as revenue,
               SUM(profit) as profit
        FROM deals
        WHERE deal_type = 'sell' 
          AND date(deal_date) = ?
          AND status = 'completed'
    ''', (date,))
    sales = c.fetchone()
    report['sales'] = {
        'count': sales['count'] or 0,
        'revenue': sales['revenue'] or 0,
        'profit': sales['profit'] or 0
    }
    
    # 2. 進貨統計
    c.execute('''
        SELECT COUNT(*) as count, SUM(amount) as total
        FROM deals
        WHERE deal_type = 'buy' 
          AND date(deal_date) = ?
          AND status = 'completed'
    ''', (date,))
    purchases = c.fetchone()
    report['purchases'] = {
        'count': purchases['count'] or 0,
        'total': purchases['total'] or 0
    }
    
    # 3. 新客戶
    c.execute('''
        SELECT COUNT(*) as count
        FROM customers
        WHERE date(created_at) = ?
    ''', (date,))
    report['new_customers'] = c.fetchone()['count'] or 0
    
    # 4. 客戶互動
    c.execute('''
        SELECT log_type, COUNT(*) as count
        FROM customer_logs
        WHERE date(created_at) = ?
        GROUP BY log_type
    ''', (date,))
    interactions = {row['log_type']: row['count'] for row in c.fetchall()}
    report['interactions'] = interactions
    
    # 5. 庫存狀態
    c.execute('SELECT COUNT(*) as count FROM vehicles WHERE status = "in_stock"')
    report['inventory_count'] = c.fetchone()['count'] or 0
    
    # 6. 今日成交明細
    c.execute('''
        SELECT d.id, v.brand, v.model, v.year, d.amount, d.profit,
               c.name as customer_name
        FROM deals d
        JOIN vehicles v ON d.vehicle_id = v.id
        LEFT JOIN customers c ON d.customer_id = c.id
        WHERE d.deal_type = 'sell'
          AND date(d.deal_date) = ?
          AND d.status = 'completed'
    ''', (date,))
    report['deal_details'] = [dict(row) for row in c.fetchall()]
    
    # 7. 與昨日比較
    yesterday = (datetime.strptime(date, '%Y-%m-%d') - timedelta(days=1)).strftime('%Y-%m-%d')
    c.execute('''
        SELECT SUM(amount) as revenue FROM deals
        WHERE deal_type = 'sell' AND date(deal_date) = ? AND status = 'completed'
    ''', (yesterday,))
    yesterday_revenue = c.fetchone()['revenue'] or 0
    
    current_revenue = report['sales']['revenue']
    if yesterday_revenue > 0:
        report['vs_yesterday'] = {
            'revenue_change': current_revenue - yesterday_revenue,
            'revenue_change_pct': round((current_revenue - yesterday_revenue) / yesterday_revenue * 100, 1)
        }
    else:
        report['vs_yesterday'] = {
            'revenue_change': current_revenue,
            'revenue_change_pct': 0
        }
    
    conn.close()
    
    # 生成文字摘要
    report['summary'] = _generate_daily_summary(report)
    
    return report


def _generate_daily_summary(report: Dict) -> str:
    """生成日報文字摘要"""
    date = report['date']
    sales = report['sales']
    
    lines = [
        f"📊 {date} 日報",
        "",
        f"💰 銷售：{sales['count']} 台，營收 ${sales['revenue']:,}，毛利 ${sales['profit']:,}",
        f"📦 進貨：{report['purchases']['count']} 台，金額 ${report['purchases']['total']:,}",
        f"👥 新客戶：{report['new_customers']} 位",
        f"🚗 庫存：{report['inventory_count']} 台",
    ]
    
    vs = report.get('vs_yesterday', {})
    if vs.get('revenue_change_pct'):
        emoji = "📈" if vs['revenue_change_pct'] > 0 else "📉"
        lines.append(f"{emoji} 較昨日：{vs['revenue_change_pct']:+.1f}%")
    
    return "\n".join(lines)


# ============================================================
# 2. 週報
# ============================================================

def generate_weekly_report(db_path: str, end_date: Optional[str] = None) -> Dict:
    """生成週報
    
    Args:
        db_path: 資料庫路徑
        end_date: 週末日期，預設本週日
    """
    if not end_date:
        today = datetime.now()
        # 找到本週日
        days_until_sunday = 6 - today.weekday()
        end_date = (today + timedelta(days=days_until_sunday)).strftime('%Y-%m-%d')
    
    end_dt = datetime.strptime(end_date, '%Y-%m-%d')
    start_date = (end_dt - timedelta(days=6)).strftime('%Y-%m-%d')
    
    conn = get_connection(db_path)
    c = conn.cursor()
    
    report = {
        'type': 'weekly',
        'start_date': start_date,
        'end_date': end_date,
        'generated_at': datetime.now().isoformat()
    }
    
    # 1. 週銷售統計
    c.execute('''
        SELECT COUNT(*) as count,
               SUM(amount) as revenue,
               SUM(profit) as profit
        FROM deals
        WHERE deal_type = 'sell'
          AND deal_date BETWEEN ? AND ?
          AND status = 'completed'
    ''', (start_date, end_date))
    sales = c.fetchone()
    report['sales'] = {
        'count': sales['count'] or 0,
        'revenue': sales['revenue'] or 0,
        'profit': sales['profit'] or 0,
        'avg_per_deal': int((sales['revenue'] or 0) / max(sales['count'] or 1, 1))
    }
    
    # 2. 每日趨勢
    c.execute('''
        SELECT date(deal_date) as date,
               COUNT(*) as count,
               SUM(amount) as revenue
        FROM deals
        WHERE deal_type = 'sell'
          AND deal_date BETWEEN ? AND ?
          AND status = 'completed'
        GROUP BY date(deal_date)
        ORDER BY date
    ''', (start_date, end_date))
    daily_trend = [dict(row) for row in c.fetchall()]
    report['daily_trend'] = daily_trend
    
    # 3. 品牌分布
    c.execute('''
        SELECT v.brand, COUNT(*) as count, SUM(d.amount) as revenue
        FROM deals d
        JOIN vehicles v ON d.vehicle_id = v.id
        WHERE d.deal_type = 'sell'
          AND d.deal_date BETWEEN ? AND ?
          AND d.status = 'completed'
        GROUP BY v.brand
        ORDER BY count DESC
    ''', (start_date, end_date))
    report['brand_distribution'] = [dict(row) for row in c.fetchall()]
    
    # 4. 新客戶統計
    c.execute('''
        SELECT COUNT(*) as count
        FROM customers
        WHERE created_at BETWEEN ? AND ?
    ''', (start_date, end_date + ' 23:59:59'))
    report['new_customers'] = c.fetchone()['count'] or 0
    
    # 5. 客戶來源分析
    c.execute('''
        SELECT source, COUNT(*) as count
        FROM customers
        WHERE created_at BETWEEN ? AND ?
        GROUP BY source
        ORDER BY count DESC
    ''', (start_date, end_date + ' 23:59:59'))
    report['customer_sources'] = [dict(row) for row in c.fetchall()]
    
    # 6. 與上週比較
    prev_end = (end_dt - timedelta(days=7)).strftime('%Y-%m-%d')
    prev_start = (end_dt - timedelta(days=13)).strftime('%Y-%m-%d')
    c.execute('''
        SELECT SUM(amount) as revenue, SUM(profit) as profit, COUNT(*) as count
        FROM deals
        WHERE deal_type = 'sell'
          AND deal_date BETWEEN ? AND ?
          AND status = 'completed'
    ''', (prev_start, prev_end))
    prev = c.fetchone()
    prev_revenue = prev['revenue'] or 0
    
    current_revenue = report['sales']['revenue']
    if prev_revenue > 0:
        report['vs_last_week'] = {
            'revenue_change': current_revenue - prev_revenue,
            'revenue_change_pct': round((current_revenue - prev_revenue) / prev_revenue * 100, 1),
            'count_change': (sales['count'] or 0) - (prev['count'] or 0)
        }
    else:
        report['vs_last_week'] = {
            'revenue_change': current_revenue,
            'revenue_change_pct': 0,
            'count_change': sales['count'] or 0
        }
    
    # 7. Top 5 成交
    c.execute('''
        SELECT v.brand, v.model, v.year, d.amount, d.profit
        FROM deals d
        JOIN vehicles v ON d.vehicle_id = v.id
        WHERE d.deal_type = 'sell'
          AND d.deal_date BETWEEN ? AND ?
          AND d.status = 'completed'
        ORDER BY d.amount DESC
        LIMIT 5
    ''', (start_date, end_date))
    report['top_deals'] = [dict(row) for row in c.fetchall()]
    
    conn.close()
    
    report['summary'] = _generate_weekly_summary(report)
    
    return report


def _generate_weekly_summary(report: Dict) -> str:
    """生成週報文字摘要"""
    sales = report['sales']
    vs = report.get('vs_last_week', {})
    
    lines = [
        f"📊 週報 ({report['start_date']} ~ {report['end_date']})",
        "",
        f"💰 銷售：{sales['count']} 台",
        f"   營收：${sales['revenue']:,}",
        f"   毛利：${sales['profit']:,}",
        f"   平均單價：${sales['avg_per_deal']:,}",
        "",
        f"👥 新客戶：{report['new_customers']} 位",
    ]
    
    if vs.get('revenue_change_pct'):
        emoji = "📈" if vs['revenue_change_pct'] > 0 else "📉"
        lines.append(f"{emoji} 較上週：{vs['revenue_change_pct']:+.1f}%")
    
    if report.get('brand_distribution'):
        lines.append("")
        lines.append("🏆 品牌排行：")
        for i, b in enumerate(report['brand_distribution'][:3], 1):
            lines.append(f"   {i}. {b['brand']}：{b['count']} 台")
    
    return "\n".join(lines)


# ============================================================
# 3. 月報
# ============================================================

def generate_monthly_report(db_path: str, year_month: Optional[str] = None) -> Dict:
    """生成月報
    
    Args:
        db_path: 資料庫路徑
        year_month: 年月 (YYYY-MM)，預設本月
    """
    if not year_month:
        year_month = datetime.now().strftime('%Y-%m')
    
    conn = get_connection(db_path)
    c = conn.cursor()
    
    report = {
        'type': 'monthly',
        'year_month': year_month,
        'generated_at': datetime.now().isoformat()
    }
    
    # 1. 月銷售統計
    c.execute('''
        SELECT COUNT(*) as count,
               SUM(amount) as revenue,
               SUM(profit) as profit
        FROM deals
        WHERE deal_type = 'sell'
          AND strftime('%Y-%m', deal_date) = ?
          AND status = 'completed'
    ''', (year_month,))
    sales = c.fetchone()
    report['sales'] = {
        'count': sales['count'] or 0,
        'revenue': sales['revenue'] or 0,
        'profit': sales['profit'] or 0,
        'profit_margin': round((sales['profit'] or 0) / max(sales['revenue'] or 1, 1) * 100, 1)
    }
    
    # 2. 進貨統計
    c.execute('''
        SELECT COUNT(*) as count, SUM(amount) as total
        FROM deals
        WHERE deal_type = 'buy'
          AND strftime('%Y-%m', deal_date) = ?
          AND status = 'completed'
    ''', (year_month,))
    purchases = c.fetchone()
    report['purchases'] = {
        'count': purchases['count'] or 0,
        'total': purchases['total'] or 0
    }
    
    # 3. 週別趨勢
    c.execute('''
        SELECT strftime('%W', deal_date) as week,
               COUNT(*) as count,
               SUM(amount) as revenue,
               SUM(profit) as profit
        FROM deals
        WHERE deal_type = 'sell'
          AND strftime('%Y-%m', deal_date) = ?
          AND status = 'completed'
        GROUP BY week
        ORDER BY week
    ''', (year_month,))
    report['weekly_trend'] = [dict(row) for row in c.fetchall()]
    
    # 4. 品牌分析
    c.execute('''
        SELECT v.brand,
               COUNT(*) as count,
               SUM(d.amount) as revenue,
               SUM(d.profit) as profit,
               AVG(d.amount) as avg_price
        FROM deals d
        JOIN vehicles v ON d.vehicle_id = v.id
        WHERE d.deal_type = 'sell'
          AND strftime('%Y-%m', d.deal_date) = ?
          AND d.status = 'completed'
        GROUP BY v.brand
        ORDER BY revenue DESC
    ''', (year_month,))
    report['brand_analysis'] = [dict(row) for row in c.fetchall()]
    
    # 5. 業務員業績（如有）
    c.execute('''
        SELECT u.name,
               COUNT(*) as count,
               SUM(d.amount) as revenue,
               SUM(d.profit) as profit
        FROM deals d
        LEFT JOIN users u ON d.salesperson_id = u.id
        WHERE d.deal_type = 'sell'
          AND strftime('%Y-%m', d.deal_date) = ?
          AND d.status = 'completed'
        GROUP BY d.salesperson_id
        ORDER BY revenue DESC
    ''', (year_month,))
    report['salesperson_performance'] = [dict(row) for row in c.fetchall()]
    
    # 6. 客戶統計
    c.execute('''
        SELECT COUNT(*) as new_customers
        FROM customers
        WHERE strftime('%Y-%m', created_at) = ?
    ''', (year_month,))
    report['new_customers'] = c.fetchone()['new_customers'] or 0
    
    # 7. 庫存週轉
    c.execute('''
        SELECT AVG(julianday(deal_date) - julianday(v.purchase_date)) as avg_days
        FROM deals d
        JOIN vehicles v ON d.vehicle_id = v.id
        WHERE d.deal_type = 'sell'
          AND strftime('%Y-%m', d.deal_date) = ?
          AND d.status = 'completed'
    ''', (year_month,))
    avg_days = c.fetchone()['avg_days']
    report['inventory_turnover'] = {
        'avg_days_to_sell': round(avg_days, 1) if avg_days else 0
    }
    
    # 8. 與上月比較
    year = int(year_month[:4])
    month = int(year_month[5:])
    if month == 1:
        prev_month = f"{year-1}-12"
    else:
        prev_month = f"{year}-{month-1:02d}"
    
    c.execute('''
        SELECT SUM(amount) as revenue, SUM(profit) as profit
        FROM deals
        WHERE deal_type = 'sell'
          AND strftime('%Y-%m', deal_date) = ?
          AND status = 'completed'
    ''', (prev_month,))
    prev = c.fetchone()
    prev_revenue = prev['revenue'] or 0
    
    current_revenue = report['sales']['revenue']
    if prev_revenue > 0:
        report['vs_last_month'] = {
            'revenue_change': current_revenue - prev_revenue,
            'revenue_change_pct': round((current_revenue - prev_revenue) / prev_revenue * 100, 1)
        }
    else:
        report['vs_last_month'] = {'revenue_change': current_revenue, 'revenue_change_pct': 0}
    
    conn.close()
    
    report['summary'] = _generate_monthly_summary(report)
    
    return report


def _generate_monthly_summary(report: Dict) -> str:
    """生成月報文字摘要"""
    sales = report['sales']
    
    lines = [
        f"📊 {report['year_month']} 月報",
        "",
        f"💰 銷售業績",
        f"   成交：{sales['count']} 台",
        f"   營收：${sales['revenue']:,}",
        f"   毛利：${sales['profit']:,}",
        f"   毛利率：{sales['profit_margin']}%",
        "",
        f"📦 進貨：{report['purchases']['count']} 台，${report['purchases']['total']:,}",
        f"👥 新客戶：{report['new_customers']} 位",
        f"⏱️ 平均庫存天數：{report['inventory_turnover']['avg_days_to_sell']} 天",
    ]
    
    vs = report.get('vs_last_month', {})
    if vs.get('revenue_change_pct'):
        emoji = "📈" if vs['revenue_change_pct'] > 0 else "📉"
        lines.append(f"{emoji} 較上月：{vs['revenue_change_pct']:+.1f}%")
    
    return "\n".join(lines)


# ============================================================
# 4. 業績排行榜
# ============================================================

def get_leaderboard(db_path: str, period: str = 'month', limit: int = 10) -> Dict:
    """取得業績排行榜
    
    Args:
        db_path: 資料庫路徑
        period: 'day', 'week', 'month', 'year'
        limit: 排行數量
    """
    conn = get_connection(db_path)
    c = conn.cursor()
    
    # 決定日期範圍
    today = datetime.now()
    if period == 'day':
        start_date = today.strftime('%Y-%m-%d')
        title = f"{start_date} 日排行"
    elif period == 'week':
        start_date = (today - timedelta(days=today.weekday())).strftime('%Y-%m-%d')
        title = "本週排行"
    elif period == 'month':
        start_date = today.strftime('%Y-%m-01')
        title = f"{today.strftime('%Y年%m月')} 排行"
    else:  # year
        start_date = today.strftime('%Y-01-01')
        title = f"{today.year}年 排行"
    
    # 業務員排行
    c.execute('''
        SELECT u.id, u.name,
               COUNT(*) as deal_count,
               SUM(d.amount) as revenue,
               SUM(d.profit) as profit
        FROM deals d
        LEFT JOIN users u ON d.salesperson_id = u.id
        WHERE d.deal_type = 'sell'
          AND d.deal_date >= ?
          AND d.status = 'completed'
        GROUP BY d.salesperson_id
        ORDER BY revenue DESC
        LIMIT ?
    ''', (start_date, limit))
    
    rankings = []
    for i, row in enumerate(c.fetchall(), 1):
        rankings.append({
            'rank': i,
            'id': row['id'],
            'name': row['name'] or '未指定',
            'deal_count': row['deal_count'],
            'revenue': row['revenue'] or 0,
            'profit': row['profit'] or 0
        })
    
    # 總計
    c.execute('''
        SELECT COUNT(*) as count, SUM(amount) as revenue, SUM(profit) as profit
        FROM deals
        WHERE deal_type = 'sell' AND deal_date >= ? AND status = 'completed'
    ''', (start_date,))
    total = c.fetchone()
    
    conn.close()
    
    return {
        'success': True,
        'title': title,
        'period': period,
        'start_date': start_date,
        'rankings': rankings,
        'total': {
            'count': total['count'] or 0,
            'revenue': total['revenue'] or 0,
            'profit': total['profit'] or 0
        }
    }


# ============================================================
# 5. Excel 匯出
# ============================================================

def export_report_to_excel(report: Dict, filename: Optional[str] = None) -> bytes:
    """將報表匯出為 Excel
    
    Args:
        report: 報表數據
        filename: 檔名（可選）
    
    Returns:
        Excel 檔案的 bytes
    """
    if not HAS_OPENPYXL:
        raise ImportError("需要安裝 openpyxl: pip install openpyxl")
    
    wb = Workbook()
    ws = wb.active
    
    report_type = report.get('type', 'report')
    
    # 設定樣式
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    
    # 標題
    if report_type == 'daily':
        ws.cell(row=row, column=1, value=f"日報 - {report['date']}").font = title_font
    elif report_type == 'weekly':
        ws.cell(row=row, column=1, value=f"週報 - {report['start_date']} ~ {report['end_date']}").font = title_font
    elif report_type == 'monthly':
        ws.cell(row=row, column=1, value=f"月報 - {report['year_month']}").font = title_font
    
    row += 2
    
    # 銷售摘要
    ws.cell(row=row, column=1, value="銷售統計").font = header_font
    row += 1
    
    sales = report.get('sales', {})
    ws.cell(row=row, column=1, value="成交數")
    ws.cell(row=row, column=2, value=sales.get('count', 0))
    row += 1
    ws.cell(row=row, column=1, value="營收")
    ws.cell(row=row, column=2, value=sales.get('revenue', 0))
    ws.cell(row=row, column=2).number_format = '#,##0'
    row += 1
    ws.cell(row=row, column=1, value="毛利")
    ws.cell(row=row, column=2, value=sales.get('profit', 0))
    ws.cell(row=row, column=2).number_format = '#,##0'
    row += 2
    
    # 品牌分析（如果有）
    brand_data = report.get('brand_distribution') or report.get('brand_analysis')
    if brand_data:
        ws.cell(row=row, column=1, value="品牌分析").font = header_font
        row += 1
        
        headers = ['品牌', '數量', '營收']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        row += 1
        
        for b in brand_data:
            ws.cell(row=row, column=1, value=b.get('brand', '')).border = border
            ws.cell(row=row, column=2, value=b.get('count', 0)).border = border
            cell = ws.cell(row=row, column=3, value=b.get('revenue', 0))
            cell.border = border
            cell.number_format = '#,##0'
            row += 1
        
        row += 1
    
    # 業務員業績（如果有）
    sp_data = report.get('salesperson_performance')
    if sp_data:
        ws.cell(row=row, column=1, value="業務員業績").font = header_font
        row += 1
        
        headers = ['姓名', '成交數', '營收', '毛利']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        row += 1
        
        for sp in sp_data:
            ws.cell(row=row, column=1, value=sp.get('name') or '未指定').border = border
            ws.cell(row=row, column=2, value=sp.get('count', 0)).border = border
            ws.cell(row=row, column=3, value=sp.get('revenue', 0)).border = border
            ws.cell(row=row, column=4, value=sp.get('profit', 0)).border = border
            row += 1
    
    # 調整欄寬
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # 輸出
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def export_deals_to_excel(db_path: str, start_date: str, end_date: str) -> bytes:
    """匯出交易明細為 Excel"""
    if not HAS_OPENPYXL:
        raise ImportError("需要安裝 openpyxl")
    
    conn = get_connection(db_path)
    c = conn.cursor()
    
    c.execute('''
        SELECT d.id, d.deal_date, d.deal_type, 
               v.brand, v.model, v.year, v.plate_number,
               c.name as customer_name,
               d.amount, d.profit, d.status
        FROM deals d
        LEFT JOIN vehicles v ON d.vehicle_id = v.id
        LEFT JOIN customers c ON d.customer_id = c.id
        WHERE d.deal_date BETWEEN ? AND ?
        ORDER BY d.deal_date DESC
    ''', (start_date, end_date))
    
    deals = c.fetchall()
    conn.close()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "交易明細"
    
    # 表頭
    headers = ['編號', '日期', '類型', '品牌', '型號', '年份', '車牌', '客戶', '金額', '毛利', '狀態']
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    
    # 資料
    type_map = {'sell': '銷售', 'buy': '進貨'}
    status_map = {'completed': '完成', 'pending': '進行中', 'cancelled': '取消'}
    
    for row, d in enumerate(deals, 2):
        ws.cell(row=row, column=1, value=d['id'])
        ws.cell(row=row, column=2, value=d['deal_date'])
        ws.cell(row=row, column=3, value=type_map.get(d['deal_type'], d['deal_type']))
        ws.cell(row=row, column=4, value=d['brand'])
        ws.cell(row=row, column=5, value=d['model'])
        ws.cell(row=row, column=6, value=d['year'])
        ws.cell(row=row, column=7, value=d['plate_number'])
        ws.cell(row=row, column=8, value=d['customer_name'])
        ws.cell(row=row, column=9, value=d['amount']).number_format = '#,##0'
        ws.cell(row=row, column=10, value=d['profit']).number_format = '#,##0'
        ws.cell(row=row, column=11, value=status_map.get(d['status'], d['status']))
    
    # 調整欄寬
    widths = [8, 12, 8, 10, 12, 8, 10, 12, 12, 12, 8]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


# 📚 知識點
# -----------
# 1. 報表設計原則：
#    - 層次分明：日/週/月
#    - 比較維度：環比（vs上期）
#    - 多角度：銷售/進貨/客戶/庫存
#
# 2. SQL 日期函數：
#    - strftime('%Y-%m', date)：提取年月
#    - strftime('%W', date)：提取週數
#    - date('now', '-7 days')：日期運算
#
# 3. Excel 處理：
#    - openpyxl：Python Excel 庫
#    - 樣式：Font, Fill, Border
#    - 格式：number_format = '#,##0'
#
# 4. BytesIO：
#    - 記憶體中的檔案物件
#    - 避免寫入磁碟
#    - 直接返回 bytes 供下載
